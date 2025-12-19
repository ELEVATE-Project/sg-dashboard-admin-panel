import openpyxl
import os
import importlib.util
import re
import requests

from constants import (
    PAGE_METADATA,
    TABS_METADATA,
    DRIVE_IMAGE_URL,
    DRIVE_DOWNLOAD_URL,
    BUCKET_PREFIX_FOR_IMAGES as prefix
)

def convert_drive_link_to_direct_url(link):
    if not isinstance(link, str):
        return ''
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", link) or re.search(r"id=([a-zA-Z0-9_-]+)", link)
    if match:
        return f"{DRIVE_IMAGE_URL}{match.group(1)}"
    return link.strip()


def download_image(file_id, save_path):
    url = f"{DRIVE_DOWNLOAD_URL}{file_id}"
    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            return True
        return False
    except Exception as e:
        print(f"‚ùå Download error: {e}")
        return False


def upload_images_from_excel(excel_file):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        images_dir = os.path.join(script_dir, "temp_downloads")
        os.makedirs(images_dir, exist_ok=True)

        workbook = openpyxl.load_workbook(excel_file, data_only=False)

        sheet_name = PAGE_METADATA["UPLOAD_IMAGES"]
        if sheet_name not in workbook.sheetnames:
            print(f"‚ùå Sheet '{sheet_name}' not found")
            print("üìÑ Available sheets:", workbook.sheetnames)
            return

        sheet = workbook[sheet_name]

        headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
        expected_columns = TABS_METADATA["UPLOAD_IMAGES"]

        if not all(col in headers for col in expected_columns):
            print("‚ùå Missing required columns")
            print("Expected:", expected_columns)
            print("Found:", headers)
            return

        name_col = headers.index(expected_columns[0])
        src_col = headers.index(expected_columns[1])

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            try:
                raw_name = row[name_col].value or ''
                raw_src = row[src_col].value or ''

                if not raw_name or not raw_src:
                    continue

                logo_url = convert_drive_link_to_direct_url(raw_src)

                file_match = re.search(r"id=([a-zA-Z0-9_-]+)", logo_url)
                if not file_match:
                    print(f"‚ö†Ô∏è No file ID found (row {row_idx})")
                    continue

                file_id = file_match.group(1)

                name_clean = re.sub(
                    r'[^a-z0-9_-]',
                    '',
                    raw_name.lower().replace(" ", "_")
                )

                local_filename = f"{name_clean}.svg"
                local_path = os.path.join(images_dir, local_filename)

                if not download_image(file_id, local_path):
                    print(f"‚ùå Download failed (row {row_idx})")
                    continue

                # Load GCP uploader dynamically
                gcp_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
                spec = importlib.util.spec_from_file_location("gcp_access", gcp_path)
                gcp_access = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(gcp_access)

                folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                    bucket_name=os.environ.get("BUCKET_NAME"),
                    source_file_path=local_path,
                    destination_blob_name=f"{prefix}{local_filename}"
                )

                if folder_url:
                    os.remove(local_path)
                    final_src = f"{folder_url.rstrip('/')}/{local_filename}"
                    print(f"‚úÖ Uploaded (row {row_idx}): {final_src}")
                else:
                    print(f"‚ùå GCS upload failed (row {row_idx})")

            except Exception as row_error:
                print(f"‚ùå Row {row_idx} error: {row_error}")

    except Exception as e:
        print(f"‚ùå Fatal error: {e}")
