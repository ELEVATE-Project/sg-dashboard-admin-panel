import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA, TABS_METADATA


def update_voices_json(excel_file):

    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path_voice = os.path.join(script_dir, "..", "pages", "voices-from-the-ground.json")

    # Load Excel file
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    try:
        sheet = workbook["Micro improvements progress"]
    except KeyError:
        print("‚ùå Sheet 'Micro improvements progress' not found.")
        return

    # Sum only 2025 data (ignore district rows)
    sums_2025 = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    valid_quarters_2025 = {'Q1': False, 'Q2': False, 'Q3': False, 'Q4': False}

    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        district_name = row[1]
        year = row[2]
        q1, q2, q3, q4 = row[3:7]

        if district_name:  # skip district rows
            continue
        if year != 2025:
            continue

        if q1 is not None:
            sums_2025['Q1'] += float(q1)
            valid_quarters_2025['Q1'] = True
        if q2 is not None:
            sums_2025['Q2'] += float(q2)
            valid_quarters_2025['Q2'] = True
        if q3 is not None:
            sums_2025['Q3'] += float(q3)
            valid_quarters_2025['Q3'] = True
        if q4 is not None:
            sums_2025['Q4'] += float(q4)
            valid_quarters_2025['Q4'] = True

    # Format 2025 data
    data_2025 = [sums_2025[q] for q in ['Q1', 'Q2', 'Q3', 'Q4'] if valid_quarters_2025[q]]

    # Update voices-from-the-ground.json
    try:
        with open(json_path_voice, 'r', encoding='utf-8') as f:
            voices_data = json.load(f)

        # Replace the 'data' field for the target type
        for item in voices_data:
            if item.get("type") == "micro-improvements-so-far":
                item["data"] = [{"year": 2025, "data": data_2025}]

        with open(json_path_voice, 'w', encoding='utf-8') as f:
            json.dump(voices_data, f, indent=2, ensure_ascii=False)

        print(f"‚úÖ Updated voices-from-the-ground.json with 2025 data: {data_2025}")

    except FileNotFoundError:
        print(f"‚ùå {json_path_voice} not found.")
        return
    except Exception as e:
        print(f"‚ùå Error updating voices-from-the-ground.json: {str(e)}")
        return

    # Upload to GCS
    try:
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path_voice,
            destination_blob_name="sg-dashboard/voices-from-the-ground.json"
        )

        if folder_url:
            print(f"‚úÖ Uploaded voices-from-the-ground.json to GCS: {folder_url}")
        else:
            print(f"‚ùå Failed to upload voices-from-the-ground.json")
    except Exception as e:
        print(f"‚ùå Error uploading to GCS: {str(e)}")




def voices_tab_big_numbers(excel_file):
    try:
        print("\nüîµ Starting Voices Tab Big Numbers Update")

        # --------------------------------------------------
        # Paths
        # --------------------------------------------------
        script_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(
            script_dir, "..", "pages", "voices-from-the-ground.json"
        )

        print(f"üìä Excel File: {excel_file}")
        print(f"üìÑ JSON Path: {json_path}")

        # --------------------------------------------------
        # Load Excel
        # --------------------------------------------------
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        sheet_name = PAGE_METADATA["VOICE_TAB_BIG_NUMBERS"]
        if sheet_name not in workbook.sheetnames:
            print("‚ùå Sheet not found:", sheet_name)
            print("Available:", workbook.sheetnames)
            return

        sheet = workbook[sheet_name]
        print(f"‚úÖ Using Excel sheet: {sheet_name}")

        # --------------------------------------------------
        # Read headers (PRESERVE INDEX)
        # --------------------------------------------------
        raw_headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [
            str(h).strip() if h is not None else ""
            for h in raw_headers
        ]

        print("\nüßæ Headers with index:")
        for i, h in enumerate(cleaned_headers):
            if h:
                print(f"   [{i}] {h}")

        # --------------------------------------------------
        # Indicator headers start from index 2
        # --------------------------------------------------
        indicator_headers = [
            str(col).strip()
            for col in TABS_METADATA["VOICE_TAB_BIG_NUMBERS"][2:]
        ]

        if not all(col in cleaned_headers for col in indicator_headers):
            print("\n‚ùå Missing required indicator columns")
            print("Expected:", indicator_headers)
            print("Found:", cleaned_headers)
            return

        print("\n‚úÖ All indicator columns are present")

        # --------------------------------------------------
        # Helper: Sum column as INTEGER
        # --------------------------------------------------
        def sum_by_header(header_name):
            idx = cleaned_headers.index(header_name)
            total = 0
            for row in sheet.iter_rows(min_row=2):
                val = row[idx].value
                if isinstance(val, (int, float)):
                    total += int(val)
            return total

        # --------------------------------------------------
        # Calculate indicator totals
        # --------------------------------------------------
        print("\nüìê Calculating column totals:")
        indicator_totals = []

        for header in indicator_headers:
            total = sum_by_header(header)
            indicator_totals.append(total)
            print(f"   ‚úî {header} ‚Üí {total}")

        # --------------------------------------------------
        # Load JSON
        # --------------------------------------------------
        with open(json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        # --------------------------------------------------
        # Update indicators
        # --------------------------------------------------
        print("\n‚úèÔ∏è Updating indicators:")
        for section in json_data:
            if section.get("type") == "data-indicators":
                for i, indicator in enumerate(section.get("indicators", [])):
                    if i >= len(indicator_totals):
                        print(f"‚ö†Ô∏è Skipping {indicator.get('label')} ‚Äî no corresponding total in Excel")
                        continue
                    old = indicator.get("value")
                    new = indicator_totals[i]
                    indicator["value"] = new
                    print(
                        f"   üîÑ {indicator.get('label')} : {old} ‚Üí {new}"
                  )

        # --------------------------------------------------
        # Write JSON
        # --------------------------------------------------
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)

        print("\nüíæ JSON updated successfully")

        # --------------------------------------------------
        # Upload to GCS
        # --------------------------------------------------
        print("\n‚òÅÔ∏è Uploading JSON to GCS...")

        gcp_access_path = os.path.join(
            script_dir, "..", "cloud-scripts", "gcp_access.py"
        )
        spec = importlib.util.spec_from_file_location(
            "gcp_access", gcp_access_path
        )
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/voices-from-the-ground.json",
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

        print("\nüéâ Voices Tab Big Numbers update completed\n")

    except Exception as e:
        print(f"\n‚ùå Error occurred: {str(e)}\n")
