from io import BytesIO
import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA


QUARTERS = ['Q1', 'Q2', 'Q3', 'Q4']


def _init_year_bucket():
    return {
        'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0,
        'valid_Q1': False, 'valid_Q2': False, 'valid_Q3': False, 'valid_Q4': False
    }


def _add_quarters(bucket, values):
    for q, v in zip(QUARTERS, values):
        if v is not None:
            bucket[q] += float(v)
            bucket[f'valid_{q}'] = True


def _format_year_data(bucket):
    return [bucket[q] for q in QUARTERS if bucket[f'valid_{q}']]


def update_voices_json_line_chart(excel_file):

    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path_voice = os.path.join(script_dir, "..", "pages", "voices-from-the-ground.json")

    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    try:
        sheet = workbook["Micro improvements progress"]
    except KeyError:
        print("❌ Sheet 'Micro improvements progress' not found.")
        return

    sums_2025 = {q: 0 for q in QUARTERS}
    valid_quarters_2025 = {q: False for q in QUARTERS}

    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        district_name = row[1]
        year = row[2]
        q_values = row[3:7]

        if district_name:
            continue
        if year != 2025:
            continue

        for q, v in zip(QUARTERS, q_values):
            if v is not None:
                sums_2025[q] += float(v)
                valid_quarters_2025[q] = True

    data_2025 = [sums_2025[q] for q in QUARTERS if valid_quarters_2025[q]]

    try:
        with open(json_path_voice, 'r', encoding='utf-8') as f:
            voices_data = json.load(f)

        for item in voices_data:
            if item.get("type") == "micro-improvements-so-far":
                item["data"] = [{"year": 2025, "data": data_2025}]

        with open(json_path_voice, 'w', encoding='utf-8') as f:
            json.dump(voices_data, f, indent=2, ensure_ascii=False)

        print(f"✅ Updated voices-from-the-ground.json with 2025 data: {data_2025}")

    except FileNotFoundError:
        print(f"❌ {json_path_voice} not found.")
        return
    except Exception as e:
        print(f"❌ Error updating voices-from-the-ground.json: {str(e)}")
        return

    try:
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path_voice,
            destination_blob_name="sg-dashboard/voices-from-the-ground.json"
        )

        if folder_url:
            print(f"✅ Uploaded voices-from-the-ground.json to GCS: {folder_url}")
        else:
            print(f"❌ Failed to upload voices-from-the-ground.json")
    except Exception as e:
        print(f"❌ Error uploading to GCS: {str(e)}")


def extract_micro_improvements(excel_file):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print(script_dir)

    json_path = os.path.join(script_dir, "..", "pages", "dashboard.json")

    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    try:
        sheet = workbook["Micro improvements progress"]
    except KeyError:
        print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
        print(f"Available sheets: {workbook.sheetnames}")
        return

    print(workbook, sheet)

    year_map = {
        2024: _init_year_bucket(),
        2025: _init_year_bucket()
    }

    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        print("function called inside", row)
        district_name = row[1]
        year = row[2]
        q_values = row[3:7]
        print("data", year, *q_values)

        if district_name:
            continue

        if not any(v is not None for v in q_values):
            continue

        if year in year_map:
            _add_quarters(year_map[year], q_values)

    result = []
    for year in [2024, 2025]:
        data = _format_year_data(year_map[year])
        if data:
            result.append({"year": year, "data": data})

    print(result)

    try:
        with open(json_path, 'r') as file:
            dashboard_data = json.load(file)

        for item in dashboard_data:
            if item.get('type') == 'line-chart':
                item['data'] = result

        with open(json_path, 'w') as file:
            json.dump(dashboard_data, file, indent=2)

        print(f"Updated dashboard.json with new line-chart data: {json.dumps(result, indent=2)}")
        extract_state_line_chart(excel_file)
        return json.dumps(dashboard_data, indent=2)

    except FileNotFoundError:
        print(f"Error: dashboard.json not found at {json_path}")
        return json.dumps(result, indent=2)
    except Exception as e:
        print(f"Error updating dashboard.json: {str(e)}")
        return json.dumps(result, indent=2)


def load_state_codes():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    state_codes_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
    if not os.path.exists(state_codes_file):
        print("❌ state_code_details.json not found.")
        return None
    with open(state_codes_file, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_district_line_chart(excel_file):
    try:
        state_codes = load_state_codes()
        if not state_codes:
            return

        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook["Micro improvements progress"]
        except KeyError:
            print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        district_files_map = {}

        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=1).value
            district_name = sheet.cell(row=row_num, column=2).value
            year = sheet.cell(row=row_num, column=3).value
            q_values = [
                sheet.cell(row=row_num, column=i).value for i in range(4, 8)
            ]

            if not state_name and not district_name:
                break

            state_name = str(state_name).strip() if state_name else ""
            district_name = str(district_name).strip() if district_name else ""

            if state_name not in state_codes:
                print(f"⚠️ State '{state_name}' not found in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            state_id = state_codes[state_name]["id"]
            district_id = state_codes[state_name].get(district_name)

            if not district_id:
                print(f"⚠️ District '{district_name}' not found for state '{state_name}' in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            if district_id not in district_files_map:
                district_files_map[district_id] = {
                    "district_name": district_name,
                    "line_chart": {
                        2024: _init_year_bucket(),
                        2025: _init_year_bucket()
                    }
                }

            if year in (2024, 2025):
                _add_quarters(district_files_map[district_id]["line_chart"][year], q_values)

            row_num += 1

        workbook.close()

        script_dir = os.path.dirname(os.path.abspath(__file__))
        for dist_id, dist_data in district_files_map.items():
            dist_dir = os.path.join(script_dir, "..", "districts", str(dist_id))
            os.makedirs(dist_dir, exist_ok=True)

            line_chart_data = []
            for year in [2024, 2025]:
                data = _format_year_data(dist_data["line_chart"][year])
                if data:
                    line_chart_data.append({"year": year, "data": data})

            line_chart_path = os.path.join(dist_dir, "line-chart.json")
            with open(line_chart_path, "w", encoding="utf-8") as f:
                json.dump({"data": line_chart_data}, f, indent=2, ensure_ascii=False)

            print(f"✅ Generated line-chart.json for district {dist_id} ({dist_data['district_name']})")

    except Exception as e:
        print(f"❌ Error: {str(e)}")


def save_and_upload_state_file(script_dir, state_id, filename, data, gcp_access):
    states_dir = os.path.join(script_dir, "..", "states", str(state_id))
    os.makedirs(states_dir, exist_ok=True)

    file_path = os.path.join(states_dir, filename)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    destination_blob_name = f"sg-dashboard/states/{state_id}/{filename}"
    folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        bucket_name=os.environ.get("BUCKET_NAME"),
        source_file_path=file_path,
        destination_blob_name=destination_blob_name
    )
    if folder_url:
        print(f"✅ Uploaded {filename} for state {state_id}: {folder_url}")
    else:
        print(f"❌ Failed to upload {filename} for state {state_id}")


def extract_state_line_chart(excel_file):
    try:
        state_codes = load_state_codes()
        if not state_codes:
            return

        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook["Micro improvements progress"]
        except KeyError:
            print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        state_line_chart_map = {}

        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=1).value
            district_name = sheet.cell(row=row_num, column=2).value
            year = sheet.cell(row=row_num, column=3).value
            q_values = [
                sheet.cell(row=row_num, column=i).value for i in range(4, 8)
            ]

            if not state_name:
                break

            if district_name:
                row_num += 1
                continue

            state_name = str(state_name).strip()
            if state_name not in state_codes:
                print(f"⚠️ State '{state_name}' not found in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            state_id = state_codes[state_name]["id"]

            if state_id not in state_line_chart_map:
                state_line_chart_map[state_id] = {
                    "state_name": state_name,
                    "line_chart": {
                        2024: _init_year_bucket(),
                        2025: _init_year_bucket()
                    }
                }

            if year in (2024, 2025):
                _add_quarters(state_line_chart_map[state_id]["line_chart"][year], q_values)

            row_num += 1

        workbook.close()

        print("✅ All line-chart.json files generated & uploaded successfully.")
        extract_district_line_chart(excel_file)

    except Exception as e:
        print(f"❌ Error: {str(e)}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python extract_line_charts.py <excel_file>")
    else:
        extract_state_line_chart(sys.argv[1])
